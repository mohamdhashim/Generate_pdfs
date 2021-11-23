from fpdf import FPDF
import arabic_reshaper #ٌReshaper to show arabic correctly in pdf
from openpyxl import load_workbook 
from urllib.parse import quote
import qrcode


'''
TO run Code Successfully: 

1.cd to the directory where requirements.txt is located.
2.activate your virtualenv.
3.run: pip install -r requirements.txt in your shell.
Note: Somethimes pip fails and u must run <pip install -r requirements.txt> again

'''
class PDF(FPDF):
    #,link, image
    def QR_image(self,QR_dir,link=None):
        
        if link:
            link = 'https://ar.wikipedia.org/wiki/' + quote(link.split('/')[-1])
            
        weight,height = 110,110
        self.add_font('Arabic_font', '', 'arabic.ttf', uni=True)
        self.set_font('Arabic_font', '', 26)
        self.set_xy(210/2-weight/2.0 ,297.0/2-height/2) #to Get image in the center of A4 = 210 x 297 mm 
        self.image(QR_dir, link = link, w=weight, h=height)
        
    def book_name(self, text):

        arabic_string = arabic_reshaper.reshape(text)
        text = arabic_string[::-1]
        
        self.set_xy(50, 200)
        self.add_font('Arabic_font', '', 'arabic.ttf', uni=True) #font file loaded from external ttf file
        self.set_font('Arabic_font', '', 26)
        self.cell(w=110.0, h=40.0, align='C', txt =text, border=0 )
        
    def author_name(self, text):
        arabic_string = arabic_reshaper.reshape(text)
        text = arabic_string[::-1]
        
        self.set_xy(50, 220)
        self.add_font('Arabic_font', '', 'arabic.ttf', uni=True)
        self.set_font('Arabic_font', '', 26)
        self.cell(w=110.0, h=40.0, align='C', txt =text, border=0 )



def reverse_digits(text):
    '''
    this function reverse numbers only in the arabic text
    ex: 'ثورة 9191' <== 'ثورة 1919'
    because when using reshape arabic text (number get reversed) 
    '''
    result = ''
    curr_digit = ''
    
    for i in text:
        if i.isdigit():
            curr_digit += i
        else:
            if curr_digit:
                result = result + curr_digit[::-1]
                curr_digit = ''
            result += i
            
    if curr_digit:
            result = result + curr_digit[::-1]
            
    return result   



      
wb = load_workbook('booklist.xlsx')
ws = wb['Sheet']  
for i, book in enumerate(ws.iter_rows()):
    pdf=PDF('P', 'mm', 'A4') #A4 portrait
    pdf.add_page()
    

    link = book[1].hyperlink.target if book[1].hyperlink else None
    
    qr_img = qrcode.make(link)
    qr_img.save("QR.png")
    pdf.QR_image('QR.png',link)
    
    #reverse_digits() because when using reshaper arabic text (number get reversed in the result) 
    pdf.book_name(reverse_digits(str(book[1].value)))
    pdf.author_name(book[2].value)
    
    try:
        pdf.output('./covers/'+str(book[1].value)+'.pdf')
    except ValueError:
        print("./covers folder must be created Manually")
        




